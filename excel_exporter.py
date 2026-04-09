from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from commit_collector import CommitInfo
from pr_collector import PRRecord

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def _autofit(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def export(commits: list[CommitInfo], pr_list: list[PRRecord], repo_name: str) -> str:
    wb = Workbook()

    # --- Лист 1: Commits ---
    ws_commits = wb.active
    ws_commits.title = "Commits"
    _apply_header(ws_commits, [
        "Author", "Email", "Date", "Commit Hash", "Message", "Lines Added", "Lines Deleted"
    ])

    for c in commits:
        ws_commits.append([
            c.author_name,
            c.author_email,
            c.date.strftime("%Y-%m-%d %H:%M") if c.date else "",
            c.commit_id[:8],
            c.message,
            c.lines_added,
            c.lines_deleted,
        ])

    _autofit(ws_commits)

    # --- Лист 2: PRs ---
    ws_prs = wb.create_sheet("PRs")
    _apply_header(ws_prs, ["Author", "Email", "PR ID", "Date"])

    for pr in pr_list:
        ws_prs.append([
            pr.author_name,
            pr.author_email,
            pr.pr_id,
            pr.date.strftime("%Y-%m-%d %H:%M") if pr.date else "",
        ])

    _autofit(ws_prs)

    filename = f"{repo_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename
